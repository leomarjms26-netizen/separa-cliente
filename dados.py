import io
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gerar Planilha por Cliente (modelo)", layout="wide")

st.title("Automatizador — gerar planilha por cliente usando um modelo (Streamlit)")

st.markdown(
    """
    **Fluxo:** envie a planilha bruta e o modelo. O app vai criar uma aba para cada cliente (coluna padrão `Cliente Processo`)
    baseada no layout do modelo e preencher os dados correspondentes.
    """
)

uploaded_raw = st.file_uploader("1) Envie a planilha BRUTA (xlsx ou csv)", type=["xlsx", "xls", "csv"])
uploaded_model = st.file_uploader("2) Envie o MODELO (xlsx) — a planilha que deve ser igualada", type=["xlsx", "xls"])

if uploaded_raw and uploaded_model:
    # lê dados brutos
    try:
        if uploaded_raw.name.lower().endswith((".xls", ".xlsx")):
            df_raw = pd.read_excel(uploaded_raw, sheet_name=0)
        else:
            df_raw = pd.read_csv(uploaded_raw)
    except Exception as e:
        st.error(f"Erro ao ler arquivo bruto: {e}")
        st.stop()

    st.success(f"Arquivo bruto carregado — {df_raw.shape[0]} linhas x {df_raw.shape[1]} colunas")
    st.dataframe(df_raw.head(10))

    # detectar colunas de cliente possíveis
    candidate_cols = [c for c in df_raw.columns if "cliente" in c.lower() or "contrato" in c.lower() or "processo" in c.lower()]
    if not candidate_cols:
        candidate_cols = list(df_raw.columns)

    cliente_col = st.selectbox("Coluna que identifica o CLIENTE (será usada para agrupar):", options=candidate_cols, index=0)

    st.markdown("**Opções de limpeza de nomes de abas (remoção de caracteres inválidos):**")
    safe_truncate = st.checkbox("Truncar nomes de abas para 31 caracteres (recomendado)", value=True)
    replace_slash = st.checkbox("Substituir barras e dois-pontos por '-' nos nomes de abas", value=True)

    # carregar template workbook
    try:
        uploaded_model.seek(0)
        wb = load_workbook(uploaded_model)
    except Exception as e:
        st.error(f"Erro ao ler o arquivo modelo com openpyxl: {e}")
        st.stop()

    st.write("Modelo carregado. Abas disponíveis no modelo:")
    st.write(wb.sheetnames)

    # escolher aba modelo (por padrão a primeira)
    model_sheet_name = st.selectbox("Escolha a aba do modelo que será duplicada para cada cliente:", options=wb.sheetnames, index=0)

    if st.button("Gerar arquivo final"):
        st.info("Processando...")

        # função utilitária para deixar nome de aba seguro
        def safe_sheet_name(name):
            if name is None:
                name = "SemNome"
            name = str(name)
            if replace_slash:
                name = re.sub(r'[\/\\\:]', '-', name)
            # openpyxl limita a 31 chars
            if safe_truncate:
                name = name[:31]
            # também evita nomes vazios
            if name.strip() == "":
                name = "SemNome"
            return name

        # encontra cabeçalho no sheet do modelo (retorna row index 1-based), tentando casar com colunas do df
        def find_header_row(ws, df_columns, max_scan_rows=20):
            df_set = set([str(c).strip() for c in df_columns])
            for r in range(1, max_scan_rows + 1):
                row_vals = [str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else "" for c in range(1, ws.max_column + 1)]
                # conta quantos headers do df aparecem nessa linha
                found = sum(1 for v in row_vals if v in df_set)
                # heurística: se encontrou pelo menos metade dos campos comuns ou >=1, considera header
                if found >= 1:
                    return r, row_vals
            return 1, None

        model_ws = wb[model_sheet_name]

        # calcula clientes únicos (mantendo ordem de aparecimento)
        df_raw[cliente_col] = df_raw[cliente_col].astype(str).fillna("Sem Cliente")
        clientes = df_raw[cliente_col].astype(str)
        unique_clients = list(dict.fromkeys(clientes.tolist()))  # mantém ordem

        st.write(f"{len(unique_clients)} clientes detectados (inclui 'Sem Cliente' se aplicável).")

        # localizar header no template
        header_row_idx, header_row_vals = find_header_row(model_ws, df_raw.columns, max_scan_rows=20)
        st.write(f"Cabeçalho detectado na linha: {header_row_idx}")

        # vamos criar uma cópia para cada cliente
        created_sheets = []
        for cliente in unique_clients:
            nome_aba = safe_sheet_name(cliente)
            # copia o template
            try:
                new_ws = wb.copy_worksheet(model_ws)
            except Exception:
                # fallback: cria nova folha em branco com mesmo nome se não puder copiar
                new_ws = wb.create_sheet(title=nome_aba)
            # renomear
            # garantir não duplicar nomes existentes
            final_name = nome_aba
            i = 1
            while final_name in wb.sheetnames:
                final_name = nome_aba[:28] + f"_{i}" if len(nome_aba) > 28 else nome_aba + f"_{i}"
                i += 1
            new_ws.title = final_name

            # extrair dados deste cliente
            df_c = df_raw[df_raw[cliente_col].astype(str) == str(cliente)].copy()

            # mapear colunas do df para colunas do template por matching do header_row_vals
            template_headers = [ (idx+1, str(v).strip()) for idx, v in enumerate(header_row_vals) ] if header_row_vals else []
            col_map = {}  # df_col -> template_col_index (1-based)
            if template_headers:
                # para cada df col, procure primeira ocorrência igual no template headers
                for df_col in df_c.columns:
                    df_col_s = str(df_col).strip()
                    match = next((idx for idx, val in template_headers if val == df_col_s), None)
                    if match is not None:
                        col_map[df_col] = match
                # colunas não encontradas serão escritas após a última coluna existente
            start_row = header_row_idx + 1

            # prepara a posição para colunas extras
            last_col = new_ws.max_column
            next_free_col = last_col + 1

            # escrever linhas do dataframe
            for r_idx, (_, row) in enumerate(df_c.iterrows(), start=start_row):
                for df_col in df_c.columns:
                    if df_col in col_map:
                        c_idx = col_map[df_col]
                    else:
                        c_idx = next_free_col
                        col_map[df_col] = c_idx
                        next_free_col += 1
                    cell = new_ws.cell(row=r_idx, column=c_idx)
                    # atribui valor. Mantemos fórmulas existentes do template (não sobrescrevemos cabeçalho)
                    cell.value = row[df_col]

            created_sheets.append(final_name)

        # opcional: remover a aba modelo original se o usuário quiser (mas aqui deixamos)
        # salvar em bytes para download
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        st.success("Arquivo gerado com sucesso!")
        st.write("Abas criadas:")
        st.write(created_sheets[:50])  # mostra até 50 nomes

        # nome arquivo de saída
        out_name = "Planilha_GERAL_CLIENTES_GERADA.xlsx"

        st.download_button(
            label="⤓ Baixar planilha final",
            data=bio,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.info("Se precisar que o app também copie fórmulas/formatos mais complexos do modelo, posso ajustar o script para preservar estilos por coluna/linha (alguns templates exigem regras extras).")
